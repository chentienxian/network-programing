# coding:utf-8

import openpyxl
import re

board_ge = "PGE"
board_10ge = "PXG"
board_50ge = "PHC"
board_100ge = "PCG"
board_200ge = "PDC"

pattern_board_num = r".*\[0-1-(\d+)\]*"  # board number example: PXGA12T1[0-1-1]
pattern_port_num = r".*:(\d+)$"  # port number example: 10GE:8

re_board = re.compile(pattern_board_num)
re_port = re.compile(pattern_port_num)


def reg_board_port(board, port):
    group_board_num = re_board.search(board)
    group_port_num = re_port.search(port)
    board_num = group_board_num.group(1)
    port_num = group_port_num.group(1)
    raw_str = board_num + "/0/" + port_num
    return raw_str


def convert(board, port):
    convert_port = " "
    if board_ge in board:
        raw_board_port = reg_board_port(board, port)
        convert_port = "gei-1/" + raw_board_port
    elif board_10ge in board and "10GE" in port:
        raw_board_port = reg_board_port(board, port)
        convert_port = "xgei-1/" + raw_board_port
    elif board_50ge in board and "50GE" in port:
        raw_board_port = reg_board_port(board, port)
        convert_port = "lgei-1/" + raw_board_port
    elif board_100ge in board and "100GE" in port:
        raw_board_port = reg_board_port(board, port)
        convert_port = "cgei-1/" + raw_board_port
    elif board_200ge in board_name and "200GE" in port:
        raw_board_port = reg_board_port(board, port)
        convert_port = "ccgei-1/" + raw_board_port
    return convert_port


wb = openpyxl.load_workbook("67 links 20210414.xlsx")
sheet1 = wb["sheet1"]

i = 2
for i in range(i, sheet1.max_row + 1):
    board_name = sheet1.cell(row=i, column=4).value
    port_number = sheet1.cell(row=i, column=5).value
    if board_name != None and port_number != None:
        convert_str = convert(board_name, port_number)
    else:
        pass
        # print(board_name+port_number)
    sheet1["F" + str(i)] = convert_str

wb.save("67 links 20210414.xlsx")
print("file saved. ")